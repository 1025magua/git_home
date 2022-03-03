"""
-------------------------------------------------
   File Name:excel_utils05
   Author:Lee
   date: 2021/9/25-14:53
-------------------------------------------------
"""

import openpyxl

"""
将测试数据封装成对象
"""


class CaseData:
    def __init__(self, dict_case):
        for i in dict_case.items():  # 获取字典中的每一对键值对，并且边进行遍历
            setattr(self, i[0], i[1])


class ReadExcel:
    @classmethod
    def read_data_all(cls, file_name, sheet_name):  # 读取所有内容
        wb = openpyxl.load_workbook(file_name)  # 读取工作簿
        sh = wb[sheet_name]  # 获取sheet页
        rows = list(sh.rows)  # 获取所有单元格
        allCase = []
        # 读取表头数据
        titles = []
        for cell in rows[0]:  # rows[0]代表第一行数据
            titles.append(cell.value)  # 将第一行cell的值添加到titles中
        # print(titles)  # ['case_id','case_title','interface'.....]

        # 遍历其他行数据，和表头打包，转换为字典，放到列表
        for row in rows[1:]:  # row代表除了表头的每一行数据
            data = []  # [1,'登录成功','user_login',....]
            for v in row:
                data.append(v.value)
            case_zip = dict(zip(titles, data))  # 将数据进行打包
            cd = CaseData(case_zip)  # 调用构造方法，将数据转换为对象
            allCase.append(cd)
        return allCase

    @classmethod
    def read_data_pl(cls, file_name, sheet_name, begin_row, end_row):
        wb = openpyxl.load_workbook(file_name)
        sh = wb[sheet_name]
        rows = list(sh.rows)
        allCase = []
        # 读取表头数据
        titles = []
        for cell in rows[0]:
            titles.append(cell.value)

        # 遍历其他行数据,和表头打包，转换为字典，存放到列表
        for row in rows[begin_row: end_row + 1]:
            data = []  # [1,'登录成功','user_login',....]
            for v in row:
                data.append(v.value)
            case_zip = dict(zip(titles, data))  # 将数据进行打包
            cd = CaseData(case_zip)  # 调用构造方法，将数据转换为对象
            allCase.append(cd)
        return allCase

    @classmethod
    def write_data(cls, file_name, sheet_name, row, column, value):
        """
        :param file_name: 需要读写的文件
        :param sheet_name: 需要读写的sheet页
        :param row: 需要被写入的行
        :param column: 需要被写入的列
        :param value: 需要写入的内容
        """
        wb = openpyxl.load_workbook(file_name)
        sh = wb[sheet_name]
        sh.cell(row=row + 1, column=column, value=value)
        wb.save(file_name)
